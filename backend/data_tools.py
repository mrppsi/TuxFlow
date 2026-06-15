from __future__ import annotations

import csv
import math
import statistics
from collections import Counter
from pathlib import Path
from typing import Any


Dataset = dict[str, Any]


def read_uploaded_file(path: Path) -> Dataset:
    extension = path.suffix.lower()
    if extension == ".csv":
        return _read_csv(path)
    if extension == ".xlsx":
        return _read_xlsx(path)
    if extension == ".xls":
        return _read_xls(path)
    raise ValueError("Formato no soportado. Usa CSV, XLSX o XLS.")


def clean_dataframe(
    dataset: Dataset,
    remove_duplicates: bool = True,
    remove_empty_rows: bool = True,
    null_strategy: str = "none",
) -> Dataset:
    columns = list(dataset["columns"])
    rows = [dict(row) for row in dataset["rows"]]

    if remove_empty_rows:
        rows = [row for row in rows if not _is_empty_row(row, columns)]

    if remove_duplicates:
        rows = _remove_duplicates(rows, columns)

    if null_strategy in {"mean", "median"}:
        for column in _numeric_columns(rows, columns):
            values = [_to_number(row.get(column)) for row in rows]
            numbers = [value for value in values if value is not None]
            if not numbers:
                continue
            fill_value = statistics.mean(numbers) if null_strategy == "mean" else statistics.median(numbers)
            for row in rows:
                if _is_blank(row.get(column)):
                    row[column] = fill_value

    return {"columns": columns, "rows": rows}


def dataframe_preview(dataset: Dataset, limit: int = 20) -> list[dict[str, Any]]:
    return dataset["rows"][:limit]


def dataframe_columns(dataset: Dataset) -> list[dict[str, str]]:
    rows = dataset["rows"]
    return [{"name": column, "type": _column_type(rows, column)} for column in dataset["columns"]]


def descriptive_stats(dataset: Dataset) -> dict[str, Any]:
    columns = dataset["columns"]
    rows = dataset["rows"]
    stats: dict[str, Any] = {
        "row_count": len(rows),
        "column_count": len(columns),
        "nulls_by_column": {column: _null_count(rows, column) for column in columns},
        "duplicate_rows": _duplicate_count(rows, columns),
        "numeric": {},
    }

    for column in _numeric_columns(rows, columns):
        numbers = [_to_number(row.get(column)) for row in rows]
        numbers = [value for value in numbers if value is not None]
        if not numbers:
            continue

        stats["numeric"][column] = {
            "mean": _round(statistics.mean(numbers)),
            "median": _round(statistics.median(numbers)),
            "mode": _round(_mode(numbers)),
            "min": _round(min(numbers)),
            "max": _round(max(numbers)),
            "std": _round(statistics.stdev(numbers)) if len(numbers) > 1 else 0,
        }

    return stats


def chart_payload(dataset: Dataset) -> dict[str, Any]:
    columns = dataset["columns"]
    rows = dataset["rows"]
    numeric_columns = _numeric_columns(rows, columns)
    charts: dict[str, Any] = {"numeric_columns": numeric_columns}

    if numeric_columns:
        first = numeric_columns[0]
        values = [_to_number(row.get(first)) for row in rows]
        values = [value for value in values if value is not None][:30]
        charts["first_numeric_series"] = {
            "column": first,
            "labels": [str(index + 1) for index in range(len(values))],
            "values": [_round(value) for value in values],
        }

        means = []
        labels = []
        for column in numeric_columns:
            numbers = [_to_number(row.get(column)) for row in rows]
            numbers = [value for value in numbers if value is not None]
            if numbers:
                labels.append(column)
                means.append(_round(statistics.mean(numbers)))
        charts["numeric_means"] = {"labels": labels, "values": means}

    charts["nulls_by_column"] = {
        "labels": columns,
        "values": [_null_count(rows, column) for column in columns],
    }
    return charts


def _read_csv(path: Path) -> Dataset:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        columns = [str(column).strip() for column in (reader.fieldnames or [])]
        if not columns:
            raise ValueError("El archivo no tiene encabezados.")
        rows = [_normalize_row(row, columns) for row in reader]
    return {"columns": columns, "rows": rows}


def _read_xlsx(path: Path) -> Dataset:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Falta openpyxl. Ejecuta: pip install -r requirements.txt") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    headers = next(values, None)
    if not headers:
        raise ValueError("El archivo no tiene encabezados.")

    columns = [_clean_header(value, index) for index, value in enumerate(headers)]
    rows = []
    for values_row in values:
        row = {column: _clean_value(values_row[index] if index < len(values_row) else None) for index, column in enumerate(columns)}
        rows.append(row)
    return {"columns": columns, "rows": rows}


def _read_xls(path: Path) -> Dataset:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("Falta xlrd para leer XLS. Ejecuta: pip install -r requirements.txt") from exc

    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("El archivo no tiene encabezados.")

    columns = [_clean_header(sheet.cell_value(0, index), index) for index in range(sheet.ncols)]
    rows = []
    for row_index in range(1, sheet.nrows):
        row = {column: _clean_value(sheet.cell_value(row_index, index)) for index, column in enumerate(columns)}
        rows.append(row)
    return {"columns": columns, "rows": rows}


def _normalize_row(row: dict[str, Any], columns: list[str]) -> dict[str, Any]:
    return {column: _clean_value(row.get(column)) for column in columns}


def _clean_header(value: Any, index: int) -> str:
    text = str(value).strip() if value is not None else ""
    return text or f"column_{index + 1}"


def _clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _is_empty_row(row: dict[str, Any], columns: list[str]) -> bool:
    return all(_is_blank(row.get(column)) for column in columns)


def _remove_duplicates(rows: list[dict[str, Any]], columns: list[str]) -> list[dict[str, Any]]:
    seen = set()
    unique = []
    for row in rows:
        key = tuple(_key_value(row.get(column)) for column in columns)
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def _duplicate_count(rows: list[dict[str, Any]], columns: list[str]) -> int:
    keys = [tuple(_key_value(row.get(column)) for column in columns) for row in rows]
    counts = Counter(keys)
    return sum(count - 1 for count in counts.values() if count > 1)


def _key_value(value: Any) -> str:
    return "" if _is_blank(value) else str(value).strip().lower()


def _null_count(rows: list[dict[str, Any]], column: str) -> int:
    return sum(1 for row in rows if _is_blank(row.get(column)))


def _numeric_columns(rows: list[dict[str, Any]], columns: list[str]) -> list[str]:
    result = []
    for column in columns:
        values = [row.get(column) for row in rows if not _is_blank(row.get(column))]
        if values and all(_to_number(value) is not None for value in values):
            result.append(column)
    return result


def _column_type(rows: list[dict[str, Any]], column: str) -> str:
    values = [row.get(column) for row in rows if not _is_blank(row.get(column))]
    if not values:
        return "empty"
    if all(_to_number(value) is not None for value in values):
        return "number"
    return "text"


def _to_number(value: Any) -> float | None:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else None
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def _mode(values: list[float]) -> float:
    counts = Counter(values)
    return counts.most_common(1)[0][0]


def _round(value: float | None) -> float | None:
    if value is None:
        return None
    return round(float(value), 6)
